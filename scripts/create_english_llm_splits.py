#!/usr/bin/env python3
"""Create transcript-level English LLM validation splits and diagnostics."""

from __future__ import annotations

import csv
import json
import math
import random
import re
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


LLM_DIR = Path(__file__).resolve().parents[1]
ROOT = LLM_DIR.parents[1] if LLM_DIR.parent.name == "Data" else LLM_DIR.parent
DATASET_DIR = LLM_DIR / "datasets"
SPLIT_DIR = LLM_DIR / "splits" / "english"
FIGURE_DIR = LLM_DIR / "results" / "english_split_diagnostics"
ENGLISH_WORKBOOK = ROOT / "Data" / "Transcripts" / "English" / "2024-03-04_negation_coding_bloom_choi.xlsx"

INPUT_JSONL = DATASET_DIR / "english_llm_dataset.jsonl"
REFERENCE_JSONL = DATASET_DIR / "english_human_coding_reference.jsonl"

SEED = 20260527
N_ITERATIONS = 20000

CODER_NAMES = ("coded_by_1", "coded_by_2")
SPLIT_TARGETS = {
    "dev_train": 0.20,
    "dev_check_1": 0.25,
    "dev_check_2": 0.25,
    "test_lockbox": 0.30,
}
NEITHER_SPLIT = "uncoded_by_neither"

TRANSCRIPT_SHEETS = {
    1: "Transcript - First half",
    2: "Transcript - Second half",
}


def read_jsonl(path):
    with path.open(encoding="utf-8") as handle:
        return [json.loads(line) for line in handle]


def write_jsonl(path, records):
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def header_map(row):
    return {str(name).strip(): idx for idx, name in enumerate(row) if name is not None}


def parse_childes_age(age_text):
    """Parse CHILDES age strings such as 2;04.16 or 1;07. into months."""
    age_text = clean(age_text)
    if not age_text:
        return None
    match = re.match(r"^(\d+);(\d+)(?:\.(\d+))?", str(age_text))
    if not match:
        return None
    years = int(match.group(1))
    months = int(match.group(2))
    days = int(match.group(3) or 0)
    return years * 12 + months + days / 30.4375


def read_child_ages():
    workbook = load_workbook(ENGLISH_WORKBOOK, read_only=True, data_only=True)
    ages = {}
    raw_ages = {}

    for half, sheet_name in TRANSCRIPT_SHEETS.items():
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = header_map(next(rows))

        for row in rows:
            if clean(row[headers["Speaker"]]) != "@ID:":
                continue
            transcript_id = clean(row[headers["Transcript"]])
            utterance = clean(row[headers["Utterance"]])
            if not transcript_id or not utterance:
                continue
            parts = str(utterance).split("|")
            if len(parts) < 8:
                continue
            participant = parts[2] if len(parts) > 2 else ""
            role = parts[7] if len(parts) > 7 else ""
            if participant == "CHI" or role == "Target_Child":
                age_raw = parts[3] if len(parts) > 3 else None
                age_months = parse_childes_age(age_raw)
                if age_months is not None:
                    ages[transcript_id] = age_months
                    raw_ages[transcript_id] = age_raw

    return ages, raw_ages


def is_coded(record):
    return any(record[name] != "NA" for name in CODER_NAMES)


def coding_status(record):
    coded = [record[name] for name in CODER_NAMES if record[name] != "NA"]
    if len(coded) == 0:
        return "neither"
    if len(coded) == 1:
        return "one_coder"
    return "both_coders"


def overlap_pair(record):
    if record["coded_by_1"] == "NA" or record["coded_by_2"] == "NA":
        return None
    return f"{record['coded_by_1']} + {record['coded_by_2']}"


def add_reference_labels(records, references):
    reference_by_id = {record["record_id"]: record for record in references}
    for record in records:
        ref = reference_by_id[record["record_id"]]
        record["bloom_1"] = (ref.get("coder_1") or {}).get("bloom_label")
        record["bloom_2"] = (ref.get("coder_2") or {}).get("bloom_label")


def both_coded_for_irr(record):
    return record.get("bloom_1") is not None and record.get("bloom_2") is not None


def cohen_kappa(label_pairs):
    if not label_pairs:
        return None
    n = len(label_pairs)
    agree = sum(1 for left, right in label_pairs if left == right)
    left_counts = Counter(left for left, _ in label_pairs)
    right_counts = Counter(right for _, right in label_pairs)
    labels = set(left_counts) | set(right_counts)
    expected = sum((left_counts[label] / n) * (right_counts[label] / n) for label in labels)
    observed = agree / n
    if math.isclose(1 - expected, 0):
        return 1.0 if math.isclose(observed, 1) else 0.0
    return (observed - expected) / (1 - expected)


def irr_stats(records):
    pairs = [(record["bloom_1"], record["bloom_2"]) for record in records if both_coded_for_irr(record)]
    if not pairs:
        return {
            "n_overlap_bloom": 0,
            "bloom_exact_agreement": None,
            "bloom_cohen_kappa": None,
        }
    agreement = sum(1 for left, right in pairs if left == right) / len(pairs)
    return {
        "n_overlap_bloom": len(pairs),
        "bloom_exact_agreement": agreement,
        "bloom_cohen_kappa": cohen_kappa(pairs),
    }


def attach_age(records, age_by_transcript, raw_age_by_transcript):
    for record in records:
        transcript_id = record["transcript_id"]
        record["child_age_months"] = age_by_transcript.get(transcript_id)
        record["child_age_raw"] = raw_age_by_transcript.get(transcript_id)


def transcript_table(coded_records):
    rows = []
    grouped = defaultdict(list)
    for record in coded_records:
        grouped[record["transcript_id"]].append(record)

    for transcript_id, items in grouped.items():
        age_values = [r["child_age_months"] for r in items if r["child_age_months"] is not None]
        age = age_values[0] if age_values else None
        status_counts = Counter(coding_status(r) for r in items)
        irr = irr_stats(items)
        rows.append(
            {
                "transcript_id": transcript_id,
                "n_rows": len(items),
                "age_months": age,
                "both_coders": status_counts["both_coders"],
                "one_coder": status_counts["one_coder"],
                "n_overlap_bloom": irr["n_overlap_bloom"],
                "n_bloom_agree": sum(
                    1 for r in items if both_coded_for_irr(r) and r["bloom_1"] == r["bloom_2"]
                ),
                "bloom_1_counts": Counter(
                    r["bloom_1"] for r in items if both_coded_for_irr(r)
                ),
                "bloom_2_counts": Counter(
                    r["bloom_2"] for r in items if both_coded_for_irr(r)
                ),
                "bloom_pair_counts": Counter(
                    (r["bloom_1"], r["bloom_2"]) for r in items if both_coded_for_irr(r)
                ),
            }
        )
    return pd.DataFrame(rows)


def age_bins(transcript_df):
    known = transcript_df["age_months"].dropna()
    if known.empty:
        return np.array([0, 1])
    low = math.floor(known.min())
    high = math.ceil(known.max())
    return np.linspace(low, high, 7)


def weighted_age_hist(rows, bins):
    known = rows.dropna(subset=["age_months"])
    if known.empty:
        return np.zeros(len(bins) - 1)
    hist, _ = np.histogram(known["age_months"], bins=bins, weights=known["n_rows"])
    total = hist.sum()
    return hist / total if total else hist


def split_score(assignments, transcript_df, global_hist, bins, targets):
    total_rows = transcript_df["n_rows"].sum()
    global_both_rate = transcript_df["both_coders"].sum() / total_rows
    global_one_rate = transcript_df["one_coder"].sum() / total_rows
    global_overlap = int(transcript_df["n_overlap_bloom"].sum())
    global_agreement = (
        transcript_df["n_bloom_agree"].sum() / global_overlap if global_overlap else None
    )
    global_kappa = kappa_from_transcript_rows(transcript_df)
    global_mean_age = np.average(
        transcript_df.dropna(subset=["age_months"])["age_months"],
        weights=transcript_df.dropna(subset=["age_months"])["n_rows"],
    )

    score = 0.0
    for split_name, target_share in targets.items():
        split_rows = transcript_df[transcript_df["transcript_id"].map(assignments) == split_name]
        n_rows = split_rows["n_rows"].sum()
        if n_rows == 0:
            return float("inf")

        row_share = n_rows / total_rows
        both_rate = split_rows["both_coders"].sum() / n_rows
        one_rate = split_rows["one_coder"].sum() / n_rows
        split_overlap = int(split_rows["n_overlap_bloom"].sum())

        score += 35 * (row_share - target_share) ** 2
        score += 18 * (both_rate - global_both_rate) ** 2
        score += 18 * (one_rate - global_one_rate) ** 2
        if global_agreement is not None and split_overlap:
            split_agreement = split_rows["n_bloom_agree"].sum() / split_overlap
            score += 16 * (split_agreement - global_agreement) ** 2
        else:
            score += 5
        if global_kappa is not None and split_overlap:
            split_kappa = kappa_from_transcript_rows(split_rows)
            if split_kappa is None:
                score += 5
            else:
                score += 10 * (split_kappa - global_kappa) ** 2

        known = split_rows.dropna(subset=["age_months"])
        if known.empty:
            score += 10
        else:
            mean_age = np.average(known["age_months"], weights=known["n_rows"])
            score += 0.035 * (mean_age - global_mean_age) ** 2
            score += 8 * np.square(weighted_age_hist(split_rows, bins) - global_hist).sum()

    return score


def combine_counters(values):
    total = Counter()
    for value in values:
        total.update(value)
    return total


def kappa_from_transcript_rows(rows):
    pair_counts = combine_counters(rows["bloom_pair_counts"])
    n = sum(pair_counts.values())
    if n == 0:
        return None
    left_counts = Counter()
    right_counts = Counter()
    agree = 0
    for (left, right), count in pair_counts.items():
        left_counts[left] += count
        right_counts[right] += count
        if left == right:
            agree += count
    observed = agree / n
    labels = set(left_counts) | set(right_counts)
    expected = sum((left_counts[label] / n) * (right_counts[label] / n) for label in labels)
    if math.isclose(1 - expected, 0):
        return 1.0 if math.isclose(observed, 1) else 0.0
    return (observed - expected) / (1 - expected)


def choose_splits(transcript_df):
    rng = random.Random(SEED)
    targets = SPLIT_TARGETS
    total_rows = int(transcript_df["n_rows"].sum())
    target_rows = {name: total_rows * share for name, share in targets.items()}
    bins = age_bins(transcript_df)
    global_hist = weighted_age_hist(transcript_df, bins)

    transcript_records = transcript_df.to_dict("records")
    best_assignments = None
    best_score = float("inf")

    for _ in range(N_ITERATIONS):
        shuffled = transcript_records[:]
        rng.shuffle(shuffled)
        assignments = {}
        current_rows = {name: 0 for name in targets}

        for transcript in shuffled:
            split_order = list(targets)
            rng.shuffle(split_order)
            split_name = min(
                split_order,
                key=lambda name: (
                    current_rows[name] / target_rows[name],
                    current_rows[name],
                ),
            )
            assignments[transcript["transcript_id"]] = split_name
            current_rows[split_name] += transcript["n_rows"]

        score = split_score(assignments, transcript_df, global_hist, bins, targets)
        if score < best_score:
            best_score = score
            best_assignments = assignments

    return best_assignments, best_score, bins


def summarize_split(records):
    count_by_status = Counter(coding_status(record) for record in records)
    coder_counts = Counter()
    overlap_counts = Counter()
    ages = []
    for record in records:
        for coder_field in CODER_NAMES:
            if record[coder_field] != "NA":
                coder_counts[record[coder_field]] += 1
        pair = overlap_pair(record)
        if pair:
            overlap_counts[pair] += 1
        if record.get("child_age_months") is not None:
            ages.append(record["child_age_months"])

    irr = irr_stats(records)
    return {
        "n_rows": len(records),
        "n_transcripts": len({r["transcript_id"] for r in records}),
        "mean_age_months": round(float(np.mean(ages)), 3) if ages else None,
        "min_age_months": round(float(np.min(ages)), 3) if ages else None,
        "max_age_months": round(float(np.max(ages)), 3) if ages else None,
        "both_coders": count_by_status["both_coders"],
        "one_coder": count_by_status["one_coder"],
        "neither": count_by_status["neither"],
        "n_overlap_bloom": irr["n_overlap_bloom"],
        "bloom_exact_agreement": (
            round(float(irr["bloom_exact_agreement"]), 4)
            if irr["bloom_exact_agreement"] is not None
            else None
        ),
        "bloom_cohen_kappa": (
            round(float(irr["bloom_cohen_kappa"]), 4)
            if irr["bloom_cohen_kappa"] is not None
            else None
        ),
        "coder_counts": dict(sorted(coder_counts.items())),
        "overlap_counts": dict(sorted(overlap_counts.items())),
    }


def csv_safe_record(record):
    out = {
        "record_id": record["record_id"],
        "split": record["split"],
        "language": record["language"],
        "source_row": record["source"]["source_row"],
        "transcript_id": record["transcript_id"],
        "half": record["half"],
        "line": record["line"],
        "speaker": record["speaker"],
        "child_id": record["child_id"],
        "child_age_months": record["child_age_months"],
        "child_age_raw": record["child_age_raw"],
        "target_negator": record["target_negator"],
        "target_utterance": record["target_utterance"],
        "coded_by_1": record["coded_by_1"],
        "coded_by_2": record["coded_by_2"],
        "context_before_json": json.dumps(record["context_before"], ensure_ascii=False),
        "context_after_json": json.dumps(record["context_after"], ensure_ascii=False),
    }
    return out


def write_csv(path, records):
    if not records:
        return
    fieldnames = list(csv_safe_record(records[0]).keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(csv_safe_record(record))


def svg_text(x, y, text, size=12, weight="400", fill="#111827", anchor="start"):
    escaped = (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )
    return (
        f'<text x="{x}" y="{y}" font-family="Arial, sans-serif" font-size="{size}" '
        f'font-weight="{weight}" fill="{fill}" text-anchor="{anchor}">{escaped}</text>'
    )


def render_split_dashboard(split_records, bins, output_path):
    coded_splits = list(SPLIT_TARGETS) + [NEITHER_SPLIT]
    summaries = {name: summarize_split(split_records.get(name, [])) for name in coded_splits}

    width = 1280
    row_h = 170
    top = 90
    irr_panel_top = top + row_h * len(coded_splits) + 35
    irr_panel_h = 280
    height = irr_panel_top + irr_panel_h + 70
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        svg_text(40, 42, "English LLM Split Diagnostics", 24, "700"),
        svg_text(40, 68, f"Transcript-level pseudorandom split, seed {SEED}; age in months; context split file rows shown.", 13, "400", "#4b5563"),
    ]

    colors = {
        "both": "#2563eb",
        "one": "#14b8a6",
        "neither": "#9ca3af",
        "age": "#f97316",
    }
    coder_palette = ["#1d4ed8", "#0891b2", "#7c3aed", "#c2410c", "#15803d", "#be123c", "#475569"]

    for row_i, split_name in enumerate(coded_splits):
        y = top + row_i * row_h
        records = split_records.get(split_name, [])
        summary = summaries[split_name]
        parts.append(f'<line x1="40" y1="{y - 18}" x2="{width - 40}" y2="{y - 18}" stroke="#e5e7eb"/>')
        parts.append(svg_text(40, y + 5, split_name, 17, "700"))
        parts.append(svg_text(40, y + 29, f"{summary['n_rows']} rows, {summary['n_transcripts']} transcripts", 12, "400", "#4b5563"))
        if summary["mean_age_months"] is not None:
            parts.append(svg_text(40, y + 50, f"age mean {summary['mean_age_months']:.1f}; range {summary['min_age_months']:.1f}-{summary['max_age_months']:.1f}", 12, "400", "#4b5563"))
        else:
            parts.append(svg_text(40, y + 50, "age unavailable", 12, "400", "#4b5563"))
        if summary["bloom_cohen_kappa"] is not None:
            parts.append(svg_text(40, y + 71, f"Bloom IRR: agreement {summary['bloom_exact_agreement']:.3f}; kappa {summary['bloom_cohen_kappa']:.3f}; n={summary['n_overlap_bloom']}", 12, "400", "#4b5563"))
        else:
            parts.append(svg_text(40, y + 71, "Bloom IRR: not available", 12, "400", "#4b5563"))

        bar_x = 260
        bar_y = y - 1
        bar_w = 260
        bar_h = 22
        running = bar_x
        for key, label, color in [
            ("both_coders", "both", colors["both"]),
            ("one_coder", "one", colors["one"]),
            ("neither", "neither", colors["neither"]),
        ]:
            value = summary[key]
            w = (value / summary["n_rows"] * bar_w) if summary["n_rows"] else 0
            if w > 0:
                parts.append(f'<rect x="{running}" y="{bar_y}" width="{w}" height="{bar_h}" fill="{color}"/>')
            running += w
        parts.append(f'<rect x="{bar_x}" y="{bar_y}" width="{bar_w}" height="{bar_h}" fill="none" stroke="#111827" stroke-width="0.7"/>')
        parts.append(svg_text(bar_x, y + 42, f"both {summary['both_coders']}  |  one {summary['one_coder']}  |  neither {summary['neither']}", 12, "400", "#374151"))

        age_x = 570
        age_y = y - 5
        age_w = 250
        age_h = 58
        parts.append(svg_text(age_x, y - 15, "child age distribution", 11, "700", "#374151"))
        ages = [r["child_age_months"] for r in records if r.get("child_age_months") is not None]
        if ages:
            hist, _ = np.histogram(ages, bins=bins)
            max_hist = max(hist) or 1
            bin_w = age_w / len(hist)
            for i, value in enumerate(hist):
                h = value / max_hist * age_h
                parts.append(
                    f'<rect x="{age_x + i * bin_w + 2}" y="{age_y + age_h - h}" width="{max(1, bin_w - 4)}" height="{h}" fill="{colors["age"]}"/>'
                )
            parts.append(f'<line x1="{age_x}" y1="{age_y + age_h}" x2="{age_x + age_w}" y2="{age_y + age_h}" stroke="#111827" stroke-width="0.7"/>')
            parts.append(svg_text(age_x, age_y + age_h + 16, f"{bins[0]:.0f}", 10, "400", "#4b5563"))
            parts.append(svg_text(age_x + age_w, age_y + age_h + 16, f"{bins[-1]:.0f}", 10, "400", "#4b5563", "end"))

        coder_x = 870
        parts.append(svg_text(coder_x, y - 15, "coder line counts", 11, "700", "#374151"))
        coder_counts = Counter(summary["coder_counts"])
        if coder_counts:
            max_coder = max(coder_counts.values())
            for j, (coder, value) in enumerate(coder_counts.most_common(7)):
                cy = y + j * 17
                w = value / max_coder * 150
                color = coder_palette[j % len(coder_palette)]
                parts.append(f'<rect x="{coder_x}" y="{cy - 10}" width="{w}" height="11" fill="{color}"/>')
                parts.append(svg_text(coder_x + 158, cy, f"{coder}: {value}", 11, "400", "#111827"))

        overlap_y = y + 82
        overlap_text = "; ".join(f"{pair}: {count}" for pair, count in Counter(summary["overlap_counts"]).most_common(4))
        if not overlap_text:
            overlap_text = "No overlapping coder pairs"
        parts.append(svg_text(260, overlap_y, f"overlap: {overlap_text}", 12, "400", "#374151"))

    parts.extend(render_transcript_irr_panel(split_records, coded_splits[:-1], irr_panel_top, width))
    parts.append("</svg>")
    output_path.write_text("\n".join(parts), encoding="utf-8")


def transcript_irr_points(split_records, split_names):
    points = []
    for split_name in split_names:
        by_transcript = defaultdict(list)
        for record in split_records.get(split_name, []):
            by_transcript[record["transcript_id"]].append(record)
        for transcript_id, records in by_transcript.items():
            overlap = [record for record in records if both_coded_for_irr(record)]
            if not overlap:
                continue
            agree = sum(record["bloom_1"] == record["bloom_2"] for record in overlap)
            points.append(
                {
                    "split": split_name,
                    "transcript_id": transcript_id,
                    "n_overlap": len(overlap),
                    "agreement": agree / len(overlap),
                }
            )
    return points


def render_transcript_irr_panel(split_records, split_names, top, width):
    points = transcript_irr_points(split_records, split_names)
    panel_x = 70
    panel_y = top + 50
    panel_w = width - 140
    panel_h = 190
    colors = {
        "dev_train": "#1d4ed8",
        "dev_check_1": "#0891b2",
        "dev_check_2": "#7c3aed",
        "test_lockbox": "#c2410c",
    }

    parts = [
        f'<line x1="40" y1="{top - 16}" x2="{width - 40}" y2="{top - 16}" stroke="#e5e7eb"/>',
        svg_text(40, top + 10, "Transcript-Level Bloom IRR", 18, "700"),
        svg_text(
            40,
            top + 32,
            "Each dot is one transcript with overlapping EB/WP Bloom codes; y-axis is exact agreement; larger dots have more overlap rows.",
            12,
            "400",
            "#4b5563",
        ),
    ]

    for tick, label in [(0, "0.00"), (0.25, "0.25"), (0.5, "0.50"), (0.75, "0.75"), (1.0, "1.00")]:
        y = panel_y + panel_h - tick * panel_h
        parts.append(f'<line x1="{panel_x}" y1="{y}" x2="{panel_x + panel_w}" y2="{y}" stroke="#e5e7eb" stroke-width="0.8"/>')
        parts.append(svg_text(panel_x - 10, y + 4, label, 10, "400", "#4b5563", "end"))

    parts.append(f'<line x1="{panel_x}" y1="{panel_y}" x2="{panel_x}" y2="{panel_y + panel_h}" stroke="#111827" stroke-width="0.8"/>')
    parts.append(f'<line x1="{panel_x}" y1="{panel_y + panel_h}" x2="{panel_x + panel_w}" y2="{panel_y + panel_h}" stroke="#111827" stroke-width="0.8"/>')

    split_w = panel_w / len(split_names)
    max_overlap = max((point["n_overlap"] for point in points), default=1)
    for i, split_name in enumerate(split_names):
        cx = panel_x + i * split_w + split_w / 2
        parts.append(svg_text(cx, panel_y + panel_h + 24, split_name, 11, "700", "#374151", "middle"))
        parts.append(f'<line x1="{panel_x + i * split_w}" y1="{panel_y}" x2="{panel_x + i * split_w}" y2="{panel_y + panel_h}" stroke="#f3f4f6" stroke-width="0.8"/>')

    for point in points:
        i = split_names.index(point["split"])
        jitter_seed = sum(ord(char) for char in point["transcript_id"]) % 1000
        jitter = (jitter_seed / 999 - 0.5) * split_w * 0.68
        x = panel_x + i * split_w + split_w / 2 + jitter
        y = panel_y + panel_h - point["agreement"] * panel_h
        radius = 2.2 + 5.8 * math.sqrt(point["n_overlap"] / max_overlap)
        parts.append(
            f'<circle cx="{x:.2f}" cy="{y:.2f}" r="{radius:.2f}" fill="{colors[point["split"]]}" fill-opacity="0.48" stroke="#111827" stroke-opacity="0.25" stroke-width="0.5">'
            f'<title>{point["split"]} | {point["transcript_id"]} | agreement {point["agreement"]:.3f} | n={point["n_overlap"]}</title></circle>'
        )

    legend_x = panel_x + panel_w - 310
    legend_y = top + 7
    for j, split_name in enumerate(split_names):
        lx = legend_x + (j % 2) * 155
        ly = legend_y + (j // 2) * 18
        parts.append(f'<circle cx="{lx}" cy="{ly}" r="5" fill="{colors[split_name]}" fill-opacity="0.65"/>')
        parts.append(svg_text(lx + 10, ly + 4, split_name, 11, "400", "#374151"))

    return parts


def main():
    SPLIT_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)

    records = read_jsonl(INPUT_JSONL)
    references = read_jsonl(REFERENCE_JSONL)
    reference_by_id = {record["record_id"]: record for record in references}
    add_reference_labels(records, references)

    age_by_transcript, raw_age_by_transcript = read_child_ages()
    attach_age(records, age_by_transcript, raw_age_by_transcript)

    coded_records = [record for record in records if is_coded(record)]
    neither_records = [record for record in records if not is_coded(record)]

    tx_df = transcript_table(coded_records)
    assignments, score, bins = choose_splits(tx_df)

    split_records = {name: [] for name in list(SPLIT_TARGETS) + [NEITHER_SPLIT]}
    for record in coded_records:
        split_name = assignments[record["transcript_id"]]
        record["split"] = split_name
        split_records[split_name].append(record)
    for record in neither_records:
        record["split"] = NEITHER_SPLIT
        split_records[NEITHER_SPLIT].append(record)

    for split_name, split_items in split_records.items():
        write_jsonl(SPLIT_DIR / f"{split_name}.jsonl", split_items)
        write_csv(SPLIT_DIR / f"{split_name}.csv", split_items)

        ref_items = [reference_by_id[item["record_id"]] | {"split": split_name} for item in split_items]
        write_jsonl(SPLIT_DIR / f"{split_name}_human_reference.jsonl", ref_items)

    manifest_rows = []
    for transcript_id, split_name in sorted(assignments.items()):
        row = tx_df[tx_df["transcript_id"] == transcript_id].iloc[0].to_dict()
        row["split"] = split_name
        manifest_rows.append(row)
    manifest_df = pd.DataFrame(manifest_rows)
    manifest_df.to_csv(SPLIT_DIR / "transcript_split_manifest.csv", index=False)

    summary = {
        "seed": SEED,
        "n_iterations": N_ITERATIONS,
        "split_score": score,
        "split_targets": SPLIT_TARGETS,
        "age_bins_months": [float(x) for x in bins],
        "n_input_rows": len(records),
        "n_coded_by_at_least_one": len(coded_records),
        "n_uncoded_by_neither": len(neither_records),
        "splits": {name: summarize_split(items) for name, items in split_records.items()},
        "outputs": {
            "split_dir": str(SPLIT_DIR.relative_to(ROOT)),
            "diagnostic_svg": str((FIGURE_DIR / "english_split_diagnostics.svg").relative_to(ROOT)),
            "transcript_manifest": str((SPLIT_DIR / "transcript_split_manifest.csv").relative_to(ROOT)),
        },
    }
    (SPLIT_DIR / "split_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    summary_rows = []
    for split_name, split_summary in summary["splits"].items():
        row = {k: v for k, v in split_summary.items() if k not in {"coder_counts", "overlap_counts"}}
        row["split"] = split_name
        row["coder_counts_json"] = json.dumps(split_summary["coder_counts"], ensure_ascii=False)
        row["overlap_counts_json"] = json.dumps(split_summary["overlap_counts"], ensure_ascii=False)
        summary_rows.append(row)
    pd.DataFrame(summary_rows).to_csv(SPLIT_DIR / "split_summary.csv", index=False)

    render_split_dashboard(split_records, bins, FIGURE_DIR / "english_split_diagnostics.svg")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()

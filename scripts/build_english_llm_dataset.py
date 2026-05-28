#!/usr/bin/env python3
"""Build the unsplit English LLM negation-coding dataset."""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


LLM_DIR = Path(__file__).resolve().parents[1]
ROOT = LLM_DIR.parents[1] if LLM_DIR.parent.name == "Data" else LLM_DIR.parent
ENGLISH_DIR = ROOT / "Data" / "Transcripts" / "English"
OUT_DIR = LLM_DIR / "datasets"

FULL_WORKBOOK = ENGLISH_DIR / "2024-03-04_negation_coding_bloom_choi.xlsx"
EB_WORKBOOK = ENGLISH_DIR / "2024-03-04_negation_coding_bloom_choi_EB.xlsx"
WP_WORKBOOK = ENGLISH_DIR / "2024-03-04_negation_coding_bloom_choi_WP.xlsx"

EXCLUSION_VALUE = "RED: INCLUDED AS TT5 FOR SYNTACTIC CODING"
CONTEXT_SIZE = 20

CODE_SHEET = "Code"
TRANSCRIPT_SHEETS = {
    1: "Transcript - First half",
    2: "Transcript - Second half",
}

OUTPUT_JSONL = OUT_DIR / "english_llm_dataset.jsonl"
OUTPUT_CSV = OUT_DIR / "english_llm_dataset_flat.csv"
REFERENCE_JSONL = OUT_DIR / "english_human_coding_reference.jsonl"
SUMMARY_JSON = OUT_DIR / "english_llm_dataset_summary.json"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def header_map(row):
    return {str(name).strip(): idx for idx, name in enumerate(row) if name is not None}


def get(row, headers, name):
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return None
    return clean(row[idx])


def normalize_flag(value):
    value = clean(value)
    if value is None:
        return None
    text = str(value).strip()
    if text.lower() in {"yes", "y", "true", "1"}:
        return "Yes"
    if text.lower() in {"no", "n", "false", "0"}:
        return "No"
    return text


def read_code_sheet(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[CODE_SHEET]
    rows = sheet.iter_rows(values_only=True)
    headers = header_map(next(rows))
    records = []
    for source_row, row in enumerate(rows, start=2):
        records.append((source_row, row, headers))
    return records


def read_coder_rows(path):
    records = read_code_sheet(path)
    by_source_row = {}
    for source_row, row, headers in records:
        by_source_row[source_row] = {
            "source_row": source_row,
            "key": row_key(row, headers),
            "bloom_label": get(row, headers, "Bloom"),
            "certain_bloom": get(row, headers, "Certain.-.Bloom"),
            "other_possibility_bloom": get(row, headers, "Other.Possibility.-.Bloom"),
            "definitely_one_of_two_bloom": get(
                row, headers, "Is.it.definitely.one.of.those.two?.-.Bloom"
            ),
            "choi_label": get(row, headers, "Choi"),
            "certain_choi": get(row, headers, "Certain.-.Choi"),
            "other_possibility_choi": get(row, headers, "Other.Possibility.-.Choi"),
            "definitely_one_of_two_choi": get(
                row, headers, "Is.it.definitely.one.of.those.two?.-.Choi"
            ),
            "comments": get(row, headers, "Comments"),
            "coded_by_raw": get(row, headers, "coded_by"),
            "flags": {
                "not_a_negation": normalize_flag(get(row, headers, "Not.a.negation?")),
                "foreign_language_negation": normalize_flag(
                    get(row, headers, "Foreign.Language.Negation?")
                ),
                "mimicry": normalize_flag(get(row, headers, "Mimicry?")),
                "singing": normalize_flag(get(row, headers, "Singing?")),
                "tag_question": normalize_flag(get(row, headers, "Tag.Question?")),
                "repetition": normalize_flag(get(row, headers, "Repetition?")),
            },
        }
    return by_source_row


def row_key(row, headers):
    return (
        get(row, headers, "Transcript"),
        get(row, headers, "Negation"),
        get(row, headers, "Half"),
        get(row, headers, "Line"),
        get(row, headers, "Speaker"),
        get(row, headers, "Utterance"),
    )


def assigned_coder(coder_record, sheet_owner):
    if coder_record["bloom_label"] is None:
        return "NA"
    return coder_record["coded_by_raw"] or sheet_owner


def read_speaker_context():
    workbook = load_workbook(FULL_WORKBOOK, read_only=True, data_only=True)
    by_transcript = defaultdict(list)
    by_line = {}

    for half, sheet_name in TRANSCRIPT_SHEETS.items():
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = header_map(next(rows))

        for row in rows:
            speaker = get(row, headers, "Speaker")
            if not speaker or not str(speaker).startswith("*"):
                continue

            transcript = get(row, headers, "Transcript")
            line = get(row, headers, "Line")
            item = {
                "line": line,
                "speaker": speaker,
                "utterance": get(row, headers, "Utterance"),
                "transcript_line": get(row, headers, "transcript_line"),
            }
            by_transcript[(half, transcript)].append(item)
            by_line[(half, transcript, line)] = len(by_transcript[(half, transcript)]) - 1

    return by_transcript, by_line


def context_window(by_transcript, by_line, half, transcript, line):
    key = (half, transcript)
    index = by_line.get((half, transcript, line))
    if index is None:
        return [], []

    lines = by_transcript[key]
    before = lines[max(0, index - CONTEXT_SIZE) : index]
    after = lines[index + 1 : index + 1 + CONTEXT_SIZE]
    return before, after


def json_safe(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def write_jsonl(path, records):
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    eb_rows = read_coder_rows(EB_WORKBOOK)
    wp_rows = read_coder_rows(WP_WORKBOOK)
    context_by_transcript, context_by_line = read_speaker_context()

    llm_records = []
    reference_records = []
    alignment_mismatches = []
    context_missing = []
    dropped_red_tt5 = 0

    for source_row, row, headers in read_code_sheet(FULL_WORKBOOK):
        if get(row, headers, "exclusion") == EXCLUSION_VALUE:
            dropped_red_tt5 += 1
            continue

        base_key = row_key(row, headers)
        eb = eb_rows.get(source_row)
        wp = wp_rows.get(source_row)
        if eb is None or wp is None or eb["key"] != base_key or wp["key"] != base_key:
            alignment_mismatches.append(source_row)

        half = int(get(row, headers, "Half"))
        transcript_id = get(row, headers, "Transcript")
        line = get(row, headers, "Line")
        before, after = context_window(context_by_transcript, context_by_line, half, transcript_id, line)
        if not before and not after:
            context_missing.append(source_row)

        record_id = f"eng_{len(llm_records) + 1:06d}"
        llm_record = {
            "record_id": record_id,
            "language": "English",
            "source": {
                "workbook": str(FULL_WORKBOOK.relative_to(ROOT)),
                "sheet": CODE_SHEET,
                "source_row": source_row,
            },
            "transcript_id": transcript_id,
            "half": half,
            "line": line,
            "speaker": get(row, headers, "Speaker"),
            "child_id": get(row, headers, "Child_ID"),
            "target_negator": get(row, headers, "Negation"),
            "target_utterance": get(row, headers, "Utterance"),
            "context_window_size": CONTEXT_SIZE,
            "context_before": before,
            "context_after": after,
            "coded_by_1": assigned_coder(eb, "EB") if eb else None,
            "coded_by_2": assigned_coder(wp, "WP") if wp else None,
        }
        llm_records.append(llm_record)

        reference_records.append(
            {
                "record_id": record_id,
                "source_row": source_row,
                "coder_1_sheet": "EB",
                "coder_1": eb,
                "coder_2_sheet": "WP",
                "coder_2": wp,
            }
        )

    write_jsonl(OUTPUT_JSONL, llm_records)
    write_jsonl(REFERENCE_JSONL, reference_records)

    flat_columns = [
        "record_id",
        "language",
        "source_row",
        "transcript_id",
        "half",
        "line",
        "speaker",
        "child_id",
        "target_negator",
        "target_utterance",
        "coded_by_1",
        "coded_by_2",
        "context_before_json",
        "context_after_json",
    ]
    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=flat_columns)
        writer.writeheader()
        for record in llm_records:
            writer.writerow(
                {
                    "record_id": record["record_id"],
                    "language": record["language"],
                    "source_row": record["source"]["source_row"],
                    "transcript_id": record["transcript_id"],
                    "half": record["half"],
                    "line": record["line"],
                    "speaker": record["speaker"],
                    "child_id": record["child_id"],
                    "target_negator": record["target_negator"],
                    "target_utterance": record["target_utterance"],
                    "coded_by_1": record["coded_by_1"],
                    "coded_by_2": record["coded_by_2"],
                    "context_before_json": json.dumps(record["context_before"], ensure_ascii=False),
                    "context_after_json": json.dumps(record["context_after"], ensure_ascii=False),
                }
            )

    summary = {
        "input_workbook": str(FULL_WORKBOOK.relative_to(ROOT)),
        "coder_1_workbook": str(EB_WORKBOOK.relative_to(ROOT)),
        "coder_2_workbook": str(WP_WORKBOOK.relative_to(ROOT)),
        "context_window_size": CONTEXT_SIZE,
        "speaker_line_rule": "Speaker value starts with '*'",
        "excluded_before_output": {
            "column": "exclusion",
            "value": EXCLUSION_VALUE,
            "dropped_rows": dropped_red_tt5,
        },
        "output_rows": len(llm_records),
        "alignment_mismatch_source_rows": alignment_mismatches,
        "context_missing_source_rows": context_missing,
        "outputs": {
            "llm_jsonl": str(OUTPUT_JSONL.relative_to(ROOT)),
            "flat_csv": str(OUTPUT_CSV.relative_to(ROOT)),
            "human_reference_jsonl": str(REFERENCE_JSONL.relative_to(ROOT)),
        },
    }
    SUMMARY_JSON.write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Build the unsplit LLM negation-coding dataset for one or more languages.

Generalization of ``build_english_llm_dataset.py`` to German, Hebrew, and
Spanish. Per-language quirks (single vs split transcript sheets, missing
``Half``/``exclusion``/``Child_ID`` columns, spaced vs dotted coder headers,
Bloom-label spelling variants) are handled here and configured in
``_pipeline_config.py``.

Usage:
    python3 build_llm_dataset.py                 # builds DEFAULT_LANGUAGES
    python3 build_llm_dataset.py german hebrew   # builds the named languages
"""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter, defaultdict

from openpyxl import load_workbook

from _pipeline_config import (
    BLOOM_CANON,
    CONTEXT_SIZE,
    DATASET_DIR,
    DEFAULT_LANGUAGES,
    LANGUAGES,
    ROOT,
)

CODE_SHEET = "Code"

# Flag columns to carry from each coder sheet (canonical header -> output key).
FLAG_COLUMNS = {
    "not a negation?": "not_a_negation",
    "foreign language negation?": "foreign_language_negation",
    "mimicry?": "mimicry",
    "singing?": "singing",
    "tag question?": "tag_question",
    "repetition?": "repetition",
}


def canon(header):
    """Normalize a header so 'Not.a.negation?' == 'Not a negation?'."""
    if header is None:
        return None
    return re.sub(r"[.\s]+", " ", str(header).strip().lower()).strip()


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def header_map(row):
    out = {}
    for idx, name in enumerate(row):
        key = canon(name)
        if key and key not in out:
            out[key] = idx
    return out


def get(row, headers, *names):
    """Fetch the first non-empty value among canonical header `names`."""
    for name in names:
        idx = headers.get(canon(name))
        if idx is not None and idx < len(row):
            value = clean(row[idx])
            if value is not None:
                return value
    return None


def normalize_line(value):
    """Lines are sometimes stored as floats ('145.0'); render them as '145'."""
    value = clean(value)
    if value is None:
        return None
    text = str(value).strip()
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except (TypeError, ValueError):
        pass
    return text


def normalize_flag(value):
    value = clean(value)
    if value is None:
        return None
    text = str(value).strip()
    if text.lower() in {"yes", "y", "true", "1"}:
        return "Yes"
    if text.lower() in {"no", "n", "false", "0"}:
        return "No"
    return text


def normalize_bloom(value, unmapped):
    value = clean(value)
    if value is None:
        return None, None
    raw = str(value).strip()
    mapped = BLOOM_CANON.get(raw.lower())
    if mapped is None:
        unmapped[raw] += 1
        return raw, raw
    return mapped, raw


def read_code_sheet(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[CODE_SHEET]
    rows = sheet.iter_rows(values_only=True)
    headers = header_map(next(rows))
    records = []
    for source_row, row in enumerate(rows, start=2):
        records.append((source_row, row, headers))
    return records


def row_key(row, headers, has_half):
    return (
        get(row, headers, "Transcript"),
        get(row, headers, "Negation"),
        get(row, headers, "Half") if has_half else None,
        normalize_line(get(row, headers, "Line")),
        get(row, headers, "Speaker"),
        get(row, headers, "Utterance"),
    )


def read_coder_rows(path, initials, has_half, unmapped):
    by_source_row = {}
    for source_row, row, headers in read_code_sheet(path):
        bloom, bloom_raw = normalize_bloom(get(row, headers, "Bloom"), unmapped)
        choi = get(row, headers, "Choi")
        by_source_row[source_row] = {
            "source_row": source_row,
            "key": row_key(row, headers, has_half),
            "initials": initials,
            "bloom_label": bloom,
            "bloom_label_raw": bloom_raw,
            "certain_bloom": get(row, headers, "Certain - Bloom"),
            "other_possibility_bloom": get(row, headers, "Other Possibility - Bloom"),
            "definitely_one_of_two_bloom": get(
                row, headers, "Is it definitely one of those two? - Bloom"
            ),
            "choi_label": choi,
            "certain_choi": get(row, headers, "Certain - Choi"),
            "other_possibility_choi": get(row, headers, "Other Possibility - Choi"),
            "definitely_one_of_two_choi": get(
                row, headers, "Is it definitely one of those two? - Choi"
            ),
            "comments": get(row, headers, "Comments"),
            "coded_by_raw": get(row, headers, "coded_by"),
            "flags": {
                out_key: normalize_flag(get(row, headers, header))
                for header, out_key in FLAG_COLUMNS.items()
            },
        }
    return by_source_row


def assigned_coder(coder_record):
    if coder_record is None or coder_record["bloom_label"] is None:
        return "NA"
    return coder_record["coded_by_raw"] or coder_record["initials"]


def read_speaker_context(master_path, transcript_sheets):
    workbook = load_workbook(master_path, read_only=True, data_only=True)
    by_transcript = defaultdict(list)
    by_line = {}

    for half, sheet_name in transcript_sheets:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = header_map(next(rows))

        for row in rows:
            speaker = get(row, headers, "Speaker")
            if not speaker or not str(speaker).startswith("*"):
                continue
            transcript = get(row, headers, "Transcript")
            line = normalize_line(get(row, headers, "Line"))
            item = {
                "line": line,
                "speaker": speaker,
                "utterance": get(row, headers, "Utterance"),
            }
            by_transcript[(half, transcript)].append(item)
            by_line[(half, transcript, line)] = len(by_transcript[(half, transcript)]) - 1

    return by_transcript, by_line


def context_window(by_transcript, by_line, half, transcript, line):
    index = by_line.get((half, transcript, line))
    if index is None:
        return [], []
    lines = by_transcript[(half, transcript)]
    before = lines[max(0, index - CONTEXT_SIZE) : index]
    after = lines[index + 1 : index + 1 + CONTEXT_SIZE]
    return before, after


def write_jsonl(path, records):
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def build_language(slug):
    config = LANGUAGES[slug]
    master_path = config["master"]
    coder_specs = config["coders"]
    transcript_sheets = config["transcript_sheets"]
    has_half = any(half is not None for half, _ in transcript_sheets)
    exclusion = config["exclusion"]
    prefix = config["prefix"]

    DATASET_DIR.mkdir(parents=True, exist_ok=True)

    unmapped_bloom = Counter()
    coder_rows = [
        read_coder_rows(path, initials, has_half, unmapped_bloom)
        for initials, path in coder_specs
    ]
    context_by_transcript, context_by_line = read_speaker_context(
        master_path, transcript_sheets
    )

    # Drop pre-output exclusions (English only) before token indexing.
    kept_rows = []
    dropped_excluded = 0
    for source_row, row, headers in read_code_sheet(master_path):
        if exclusion is not None and get(row, headers, exclusion[0]) == exclusion[1]:
            dropped_excluded += 1
            continue
        kept_rows.append((source_row, row, headers))

    # One record per negator token; number tokens within an utterance so a coder
    # (human or LLM) seeing a single record knows which occurrence it is coding.
    def utterance_key(row, headers):
        return (
            get(row, headers, "Transcript"),
            get(row, headers, "Half") if has_half else None,
            normalize_line(get(row, headers, "Line")),
        )

    negators_per_utterance = Counter(
        utterance_key(row, headers) for _, row, headers in kept_rows
    )
    negators_seen = Counter()

    llm_records = []
    reference_records = []
    alignment_mismatches = []
    context_missing = []

    for source_row, row, headers in kept_rows:
        base_key = row_key(row, headers, has_half)
        coders = [rows.get(source_row) for rows in coder_rows]
        for coder in coders:
            if coder is None or coder["key"] != base_key:
                alignment_mismatches.append(source_row)
                break

        half = get(row, headers, "Half") if has_half else None
        half = int(half) if half is not None else None
        transcript_id = get(row, headers, "Transcript")
        line = normalize_line(get(row, headers, "Line"))
        before, after = context_window(
            context_by_transcript, context_by_line, half, transcript_id, line
        )
        if not before and not after:
            context_missing.append(source_row)

        token_key = (transcript_id, half, line)
        negators_seen[token_key] += 1

        record_id = f"{prefix}_{len(llm_records) + 1:06d}"
        llm_record = {
            "record_id": record_id,
            "language": config["name"],
            "source": {
                "workbook": str(master_path.relative_to(ROOT)),
                "sheet": CODE_SHEET,
                "source_row": source_row,
            },
            "transcript_id": transcript_id,
            "half": half,
            "line": line,
            "speaker": get(row, headers, "Speaker"),
            "child_id": get(row, headers, "Child_ID"),
            "target_negator": get(row, headers, "Negation"),
            "target_utterance": get(row, headers, "Utterance"),
            "negator_index_in_utterance": negators_seen[token_key],
            "negators_in_utterance": negators_per_utterance[token_key],
            "context_window_size": CONTEXT_SIZE,
            "context_before": before,
            "context_after": after,
            "coded_by_1": assigned_coder(coders[0]),
            "coded_by_2": assigned_coder(coders[1]),
        }
        llm_records.append(llm_record)

        reference_records.append(
            {
                "record_id": record_id,
                "source_row": source_row,
                "coder_1_sheet": coder_specs[0][0],
                "coder_1": coders[0],
                "coder_2_sheet": coder_specs[1][0],
                "coder_2": coders[1],
            }
        )

    out_jsonl = DATASET_DIR / f"{slug}_llm_dataset.jsonl"
    out_csv = DATASET_DIR / f"{slug}_llm_dataset_flat.csv"
    reference_jsonl = DATASET_DIR / f"{slug}_human_coding_reference.jsonl"
    summary_json = DATASET_DIR / f"{slug}_llm_dataset_summary.json"

    write_jsonl(out_jsonl, llm_records)
    write_jsonl(reference_jsonl, reference_records)

    flat_columns = [
        "record_id", "language", "source_row", "transcript_id", "half", "line",
        "speaker", "child_id", "target_negator", "target_utterance",
        "negator_index_in_utterance", "negators_in_utterance",
        "coded_by_1", "coded_by_2", "context_before_json", "context_after_json",
    ]
    with out_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=flat_columns)
        writer.writeheader()
        for record in llm_records:
            writer.writerow(
                {
                    "record_id": record["record_id"],
                    "language": record["language"],
                    "source_row": record["source"]["source_row"],
                    "transcript_id": record["transcript_id"],
                    "half": record["half"],
                    "line": record["line"],
                    "speaker": record["speaker"],
                    "child_id": record["child_id"],
                    "target_negator": record["target_negator"],
                    "target_utterance": record["target_utterance"],
                    "negator_index_in_utterance": record["negator_index_in_utterance"],
                    "negators_in_utterance": record["negators_in_utterance"],
                    "coded_by_1": record["coded_by_1"],
                    "coded_by_2": record["coded_by_2"],
                    "context_before_json": json.dumps(record["context_before"], ensure_ascii=False),
                    "context_after_json": json.dumps(record["context_after"], ensure_ascii=False),
                }
            )

    summary = {
        "language": config["name"],
        "input_workbook": str(master_path.relative_to(ROOT)),
        "coder_1_workbook": str(coder_specs[0][1].relative_to(ROOT)),
        "coder_2_workbook": str(coder_specs[1][1].relative_to(ROOT)),
        "coder_pair": [coder_specs[0][0], coder_specs[1][0]],
        "context_window_size": CONTEXT_SIZE,
        "speaker_line_rule": "Speaker value starts with '*'",
        "has_half": has_half,
        "excluded_before_output": (
            {"column": exclusion[0], "value": exclusion[1], "dropped_rows": dropped_excluded}
            if exclusion is not None
            else {"dropped_rows": 0, "note": "no exclusion column in this language"}
        ),
        "output_rows": len(llm_records),
        "n_alignment_mismatch_rows": len(alignment_mismatches),
        "alignment_mismatch_source_rows": alignment_mismatches[:50],
        "n_context_missing_rows": len(context_missing),
        "context_missing_source_rows": context_missing[:50],
        "unmapped_bloom_labels": dict(unmapped_bloom),
        "outputs": {
            "llm_jsonl": str(out_jsonl.relative_to(ROOT)),
            "flat_csv": str(out_csv.relative_to(ROOT)),
            "human_reference_jsonl": str(reference_jsonl.relative_to(ROOT)),
        },
    }
    summary_json.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    return summary


def main():
    requested = [a.lower() for a in sys.argv[1:]] or DEFAULT_LANGUAGES
    for slug in requested:
        if slug not in LANGUAGES:
            raise SystemExit(f"Unknown language '{slug}'. Known: {sorted(LANGUAGES)}")
        summary = build_language(slug)
        print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
